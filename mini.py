import cv2
import numpy as np
import openpyxl
import os
import urllib.request

selected_area = None
drawing = False
x1, y1, x2, y2 = 0, 0, 0, 0

def load_image(image_path_or_url):
    try:
        if image_path_or_url.startswith('http://') or image_path_or_url.startswith('https://'):
            resp = urllib.request.urlopen(image_path_or_url)
            image = np.asarray(bytearray(resp.read()), dtype="uint8")
            img = cv2.imdecode(image, cv2.IMREAD_COLOR)
        else:
            img = cv2.imread(image_path_or_url)

        if img is None:
            raise Exception(f"Error loading image: {image_path_or_url}")
        
        return img
    except Exception as e:
        print(f"Error loading image: {e}")
        return None

def on_mouse(event, x, y, flags, param):
    global x1, y1, x2, y2, drawing, selected_area
    
    if event == cv2.EVENT_LBUTTONDOWN:
        drawing = True
        x1, y1 = x, y
    elif event == cv2.EVENT_MOUSEMOVE:
        if drawing:
            x2, y2 = x, y
    elif event == cv2.EVENT_LBUTTONUP:
        drawing = False
        x2, y2 = x, y
        selected_area = (x1, y1, x2, y2)

def read_names_from_excel(excel_file_or_url):
    try:
        if excel_file_or_url.startswith('http://') or excel_file_or_url.startswith('https://'):
            resp = urllib.request.urlopen(excel_file_or_url)
            wb = openpyxl.load_workbook(resp)
        else:
            wb = openpyxl.load_workbook(excel_file_or_url)
        
        sheet = wb.active
        names = []
        for row in sheet.iter_rows(values_only=True):
            names.append(row[0])
        if not names:
            print(f"No names found in Excel file: {excel_file_or_url}")
        return names
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def generate_image_with_name(image, output_image, selected_area, name):
    img = image.copy()
    
    x1, y1, x2, y2 = selected_area
    
    mask = np.zeros_like(img)
    mask[y1:y2, x1:x2] = 255
    whitened_img = cv2.addWeighted(img, 1, mask, 0.5, 0)
    
    font = cv2.FONT_HERSHEY_TRIPLEX
    font_scale = 1.5
    font_thickness = 4
    
    text_size = cv2.getTextSize(name, font, font_scale, font_thickness)[0]
    
    text_x = x1 + (x2 - x1 - text_size[0]) // 2
    text_y = y1 + (y2 - y1 + text_size[1]) // 2
    
    cv2.putText(whitened_img, name, (text_x, text_y), font, font_scale, (0, 0, 0), font_thickness)
    
    cv2.imwrite(output_image, whitened_img)

if __name__ == "__main__":
    image_path_or_url = input("Enter the path or URL to the image: ")
    excel_file_or_url = input("Enter the path or URL to the Excel file: ")
    
    img = load_image(image_path_or_url)
    if img is None:
        print(f"Error loading image: {image_path_or_url}")
        exit()

    names = read_names_from_excel(excel_file_or_url)
    if not names:
        print(f"No valid names found in Excel file: {excel_file_or_url}")
        exit()
    
    num_names = min(len(names), 5)
    output_folder = 'output_images'
    os.makedirs(output_folder, exist_ok=True)

    cv2.namedWindow("Select Area")
    cv2.setMouseCallback("Select Area", on_mouse)

    while True:
        temp_img = img.copy()
        if drawing:
            cv2.rectangle(temp_img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.imshow("Select Area", temp_img)
        
        key = cv2.waitKey(1) & 0xFF
        if key == 27:
            break
        elif selected_area is not None:
            break

    cv2.destroyAllWindows()

    for i, name in enumerate(names[:num_names], start=1):
        output_image = os.path.join(output_folder, f'output_{i}.jpg')
        generate_image_with_name(img, output_image, selected_area, name)
    
    print(f"{num_names} images with names generated and saved in '{output_folder}' folder.")
